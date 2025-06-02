import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import sys
import os
import subprocess
import platform

def generate_trading_journal_excel(start_date_str, initial_capital, num_weeks, filename="trading_journal.xlsx"):
    """
    Generates an Excel file for a Trading Plan with proper formulas and formatting,
    incorporating the requested changes and fixing the NoneType split error.
    """
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    except ValueError:
        print("Error: Invalid date format. Please use YYYY-MM-DD.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active

    # Helper function to open the file
    def open_file_os_agnostic(filepath_to_open):
        """Opens the given file with the default application."""
        try:
            abs_filepath = os.path.abspath(filepath_to_open)
            print(f"\nAttempting to open: {abs_filepath}")
            if platform.system() == 'Darwin':       # macOS
                subprocess.call(('open', abs_filepath))
            elif platform.system() == 'Windows':    # Windows
                os.startfile(abs_filepath)
            else:                                   # linux variants
                subprocess.call(('xdg-open', abs_filepath))
        except FileNotFoundError:
            print(f"Error: File not found at {abs_filepath}. Cannot open.")
        except AttributeError: # os.startfile might not be available on some minimal Python installs on Windows
            if platform.system() == 'Windows':
                try: # Fallback for Windows if os.startfile is not available
                    subprocess.call(['cmd', '/c', 'start', '', abs_filepath], shell=False)
                except Exception as e_sub_win:
                    print(f"Error opening file {abs_filepath} using subprocess fallback on Windows: {e_sub_win}")
        except Exception as e_open:
            print(f"Error opening file {abs_filepath}: {e_open}")
            print("Please ensure you have a default application set for .xlsx files or that the file path is correct.")

    ws.title = "Trading Journal"
    ws.freeze_panes = "A2"

    headers = [
        "Time Slot", "Entry Time", "Exit Time", "Signal", "Status", "Momentum",
        "Order Type", "Trade Type", "Reason / Flex / Mistakes",
        "Reason for Target/Limit", "", "", "Plan Adherence", # Translated from "پایبندی پلن"
        "Stop Loss (%)", "Risk (%)", "Target Reward (R)", "Trade Duration (min)", # Changed from "Trade Duration (15 min)" for clarity
        "Max Reward (R)", "Result (R)", "Cumulative Result (R)", "Cumulative Max Reward (R)", # Column Q, R, S, T, U
        "Balance", "Max Reward Balance", "Min Balance", "Max Balance", "Min Lows % Drawdown", # Columns V, W, X, Y, Z
        "Trade Screenshot", "", "", "", "", "", "", "", "" # Columns AA onwards
    ]

    # --- Apply changes to headers and column structure ---
    # 1. Rename "Order Type" to "enter type"
    headers[6] = "Enter type"

    # 2. Remove "Reason for Target/Limit" (original index 9, column 10)
    headers.pop(9)

    # 3. Add two more empty cells for "Reason / Flex / Mistakes" merge (original index 8, column 9)
    # After pop(9), the original headers[10] and headers[11] are now at headers[9] and headers[10].
    # We need 4 empty cells after "Reason / Flex / Mistakes" (index 8).
    # So, insert two more empty strings at index 11 and 12.
    headers.insert(11, "")
    headers.insert(12, "")
    # Now, "Reason / Flex / Mistakes" (column 9) will be followed by 4 empty strings (columns 10, 11, 12, 13).
    # "Plan Adherence" will now be at column 14 (index 13).

    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, name="Dana")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)

    for col_idx, col_letter_enum in enumerate(range(1, len(headers) + 1)):
        col_letter = get_column_letter(col_letter_enum)
        current_header = headers[col_idx]
        if current_header == "Reason / Flex / Mistakes": ws.column_dimensions[col_letter].width = 10
        # Adjusted column indices for width settings
        elif col_letter_enum >= 10 and col_letter_enum <= 13 : ws.column_dimensions[col_letter].width = 10 # These are the 4 new merged empty cells
        elif current_header == "Plan Adherence": ws.column_dimensions[col_letter].width = 20
        elif col_letter_enum >= 15 and col_letter_enum <= 22: ws.column_dimensions[col_letter].width = 7.5 # Shifted by 1
        elif current_header in ["Balance", "Max Reward Balance", "Min Balance", "Max Balance", "Min Lows % Drawdown"]: ws.column_dimensions[col_letter].width = 15 # Shifted by 1 (U, V, W, X, Y)
        elif col_letter_enum >= 1 and col_letter_enum <= 8: ws.column_dimensions[col_letter].width = 11.25
        elif col_letter_enum >= 28 and col_letter_enum <= 36: ws.column_dimensions[col_letter].width = 8 # Screenshot columns AB-AJ (shifted by 1)
        else: ws.column_dimensions[col_letter].width = 15

    # Adjust merge cells for "Reason / Flex / Mistakes" and "Trade Screenshot"
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=13) # Reason / Flex / Mistakes now spans 5 columns
    ws.cell(row=1, column=9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=1, start_column=28, end_row=1, end_column=36) # Trade Screenshot shifted by 1
    ws.cell(row=1, column=28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    time_slots = ["18:30", "19:30", "20:30", "21:30"]
    current_date = start_date
    row_num = 2
    all_data_rows = []

    status_options = [
        "Range", "Trend", "Pullback", "Undefined/Transition",
        "V-Range", "V-Trend", "V-Pullback", "V-Undefined/Transition"
    ]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', allow_blank=True)
    signal_options = [
        "Weak", "Strong", "Normal",
        "V-Weak", "V-Strong", "V-Normal"
    ]
    dv_signal = DataValidation(type="list", formula1=f'"{",".join(signal_options)}"', allow_blank=True)
    momentum_options = [
        "Decreasing", "Increasing",
        "V-Decreasing", "V-Increasing"
    ]
    dv_momentum = DataValidation(type="list", formula1=f'"{",".join(momentum_options)}"', allow_blank=True)
    
    # Update dv_order_type options
    order_type_options = [
        "2R (Normal)", "2R (Mid-Candle Pullback Entry)",
        "4R (Aggressive Stop)", "4R (Mid-Candle Pullback Entry)"
    ]
    dv_order_type = DataValidation(type="list", formula1=f'"{",".join(order_type_options)}"', allow_blank=True)
    
    dv_trade_type = DataValidation(type="list", formula1='"Reversal,Continuation"', allow_blank=True)
    adherence_options = [
        "Full Adherence", "Entry Flaw", "Exit/Flexibility Flaw",
        "High Target Flaw", "Fear of entering "
    ]
    dv_adherence = DataValidation(type="list", formula1=f'"{",".join(adherence_options)}"', allow_blank=True)

    ws.add_data_validation(dv_signal)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_momentum)
    ws.add_data_validation(dv_order_type)
    ws.add_data_validation(dv_trade_type)
    ws.add_data_validation(dv_adherence)

    first_data_row = None
    last_weekly_summary_row = None

    for week in range(num_weeks):
        daily_summary_rows_info = []
        for day in range(5):
            while current_date.weekday() >= 5:
                current_date += timedelta(days=1)

            weekday_name = current_date.strftime("%A")
            day_date_str = current_date.strftime('%Y-%m-%d')
            day_date_display = f"{weekday_name} {day_date_str}"

            date_row = row_num
            # Adjusted end_column for date row merge (was 26, now 27)
            ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=27)
            date_cell = ws.cell(row=date_row, column=1)
            date_cell.value = day_date_display
            date_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Dana")
            date_cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[date_row].height = 35
            row_num += 1

            if first_data_row is None: first_data_row = row_num
            first_row_of_day_data = row_num

            for slot_time_str in time_slots:
                current_slot_start_row = row_num
                hour, minute = map(int, slot_time_str.split(":"))
                entry_options_list = [f"{(hour + (minute + 15 * i) // 60):02d}:{( (minute + 15 * i) % 60):02d}" for i in range(4)]
                exit_options_list = []
                ch, cm = hour, minute
                while ch < 24 and not (ch == 23 and cm > 30):
                    exit_options_list.append(f"{ch:02d}:{cm:02d}")
                    cm += 15
                    if cm >= 60: ch += 1; cm = 0
                
                dv_entry_for_slot = DataValidation(type="list", formula1=f'"{",".join(entry_options_list)}"', allow_blank=True)
                dv_exit_for_slot = DataValidation(type="list", formula1=f'"{",".join(exit_options_list)}"', allow_blank=True)
                ws.add_data_validation(dv_entry_for_slot)
                ws.add_data_validation(dv_exit_for_slot)

                # Define the cell address for Risk Percentage (Column P, shifted from O) for the current slot.
                risk_percentage_actual_cell = f"P{current_slot_start_row}"

                for sub_row_idx in range(2):
                    all_data_rows.append(row_num)
                    row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if row_num % 2 == 0 else None
                    for col_idx_plus_1 in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_idx_plus_1)
                        if row_fill: cell.fill = row_fill
                        current_alignment = Alignment(horizontal='center', vertical='center')
                        header_name_for_col = headers[col_idx_plus_1-1]
                        # Adjusted column index for Reason / Flex / Mistakes (was 10, now 9)
                        if header_name_for_col == "Reason / Flex / Mistakes": current_alignment.wrap_text = True
                        elif col_idx_plus_1 == 9 : current_alignment.wrap_text = True # This is the start of the merged Reason / Flex / Mistakes
                        elif header_name_for_col == "Plan Adherence": current_alignment.wrap_text = True
                        cell.alignment = current_alignment

                    if sub_row_idx == 0: ws.cell(row=row_num, column=1).value = slot_time_str
                    # Adjusted merge cells for Reason / Flex / Mistakes (was 10-12, now 9-13)
                    ws.merge_cells(start_row=row_num, start_column=9, end_row=row_num, end_column=13)
                    
                    # Adjusted column indices for data validations
                    dv_signal.add(f"D{row_num}"); dv_status.add(f"E{row_num}"); dv_momentum.add(f"F{row_num}")
                    dv_order_type.add(f"G{row_num}"); dv_trade_type.add(f"H{row_num}"); dv_adherence.add(f"N{row_num}") # M -> N
                    
                    ws.cell(row=row_num, column=16).value = 1 # P: Risk (%) (was O)
                    ws.cell(row=row_num, column=18).value = f'=IF(AND(B{row_num}<>"",C{row_num}<>""), (C{row_num}-B{row_num})*24*4, "")' # R: Trade Duration (was Q)

                    ws.cell(row=row_num, column=20).value = 0  # T: Result (R) (was S)

                    # U: Cumulative Result (was T)
                    if row_num == first_data_row: ws.cell(row=row_num, column=21).value = f"=T{row_num}"
                    elif row_num == first_row_of_day_data: ws.cell(row=row_num, column=21).value = f"=U{row_num-2}+T{row_num}"
                    else: ws.cell(row=row_num, column=21).value = f"=U{row_num-1}+T{row_num}"
                    # V: Cumulative Max Reward (was U)
                    if row_num == first_data_row: ws.cell(row=row_num, column=22).value = f"=S{row_num}"
                    elif row_num == first_row_of_day_data: ws.cell(row=row_num, column=22).value = f"=V{row_num-2}+S{row_num}"
                    else: ws.cell(row=row_num, column=22).value = f"=V{row_num-1}+S{row_num}"
                    # W: Balance (was V)
                    if row_num == first_data_row: ws.cell(row=row_num, column=23).value = f"=IF(T{row_num}=0,{initial_capital},{initial_capital}*(1+T{row_num}*P{row_num}/100))" # T, P (was S, O)
                    elif row_num == first_row_of_day_data: ws.cell(row=row_num, column=23).value = f"=IF(T{row_num}=0,W{row_num-2},W{row_num-2} + T{row_num}*{initial_capital}*{risk_percentage_actual_cell}/100)" # T, W (was S, V)
                    else: ws.cell(row=row_num, column=23).value = f"=IF(T{row_num}=0,W{row_num-1},W{row_num-1} + T{row_num}*{initial_capital}*{risk_percentage_actual_cell}/100)" # T, W (was S, V)
                    # X: Max Reward Balance (was W)
                    if row_num == first_data_row: ws.cell(row=row_num, column=24).value = f"=IF(S{row_num}=0,{initial_capital},{initial_capital}*(1+S{row_num}*P{row_num}/100))" # S, P (was R, O)
                    elif row_num == first_row_of_day_data: ws.cell(row=row_num, column=24).value = f"=IF(S{row_num}=0,X{row_num-2},X{row_num-2} + S{row_num}*{initial_capital}*{risk_percentage_actual_cell}/100)" # S, X (was R, W)
                    else: ws.cell(row=row_num, column=24).value = f"=IF(S{row_num}=0,X{row_num-1},X{row_num-1} + S{row_num}*{initial_capital}*{risk_percentage_actual_cell}/100)" # S, X (was R, W)
                    # Y: Min Balance (was X)
                    if row_num == first_data_row: ws.cell(row=row_num, column=25).value = f"=W{row_num}" # W (was V)
                    elif row_num == first_row_of_day_data: ws.cell(row=row_num, column=25).value = f"=MIN(Y{row_num-2},W{row_num})" # Y, W (was X, V)
                    else: ws.cell(row=row_num, column=25).value = f"=MIN(Y{row_num-1},W{row_num})" # Y, W (was X, V)
                    # Z: Max Balance (was Y)
                    if row_num == first_data_row: ws.cell(row=row_num, column=26).value = f"=W{row_num}" # W (was V)
                    elif row_num == first_row_of_day_data: ws.cell(row=row_num, column=26).value = f"=MAX(Z{row_num-2},W{row_num})" # Z, W (was Y, V)
                    else: ws.cell(row=row_num, column=26).value = f"=MAX(Z{row_num-1},W{row_num})" # Z, W (was Y, V)
                    # AA: Min Lows % Drawdown (was Z)
                    ws.cell(row=row_num, column=27).value = f"=IF(Y{row_num}={initial_capital},0,(Y{row_num}-{initial_capital})/{initial_capital}*100)" # Y (was X)
                    
                    dv_entry_for_slot.add(f"B{row_num}"); dv_exit_for_slot.add(f"C{row_num}")
                    ws.cell(row=row_num, column=2).number_format = "HH:mm"; ws.cell(row=row_num, column=3).number_format = "HH:mm"
                    row_num += 1

                ws.merge_cells(start_row=current_slot_start_row, start_column=1, end_row=current_slot_start_row + 1, end_column=1)
                ws.merge_cells(start_row=current_slot_start_row, start_column=4, end_row=current_slot_start_row + 1, end_column=4)
                ws.merge_cells(start_row=current_slot_start_row, start_column=5, end_row=current_slot_start_row + 1, end_column=5)
                ws.merge_cells(start_row=current_slot_start_row, start_column=6, end_row=current_slot_start_row + 1, end_column=6)
                ws.merge_cells(start_row=current_slot_start_row, start_column=16, end_row=current_slot_start_row + 1, end_column=16) # Risk (%) shifted from 15 to 16

            last_row_of_day_data = row_num -1
            if first_row_of_day_data <= last_row_of_day_data :
                 # Adjusted for new column (was 27-35, now 28-36)
                 ws.merge_cells(start_row=first_row_of_day_data, start_column=28, end_row=last_row_of_day_data, end_column=36)
            for r_idx in range(first_row_of_day_data, row_num): ws.row_dimensions[r_idx].height = 50

            summary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            # Adjusted end_column for summary merge (was 13, now 14)
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=14)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col); cell.fill = summary_fill
                cell.font = Font(bold=True, name="Dana"); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[row_num].height = 22.5
            ws.cell(row=row_num, column=1).value = "Daily Summary"
            
            if first_row_of_day_data <= last_row_of_day_data:
                ws.cell(row=row_num, column=15).value = f"=IFERROR(AVERAGE(N{first_row_of_day_data}:N{last_row_of_day_data}),\"\")" # N -> O
                ws.cell(row=row_num, column=16).value = f"=IFERROR(SUM(P{first_row_of_day_data}:P{last_row_of_day_data}),\"\")" # O -> P
                ws.cell(row=row_num, column=17).value = f"=IFERROR(SUM(Q{first_row_of_day_data}:Q{last_row_of_day_data}),\"\")" # P -> Q
                ws.cell(row=row_num, column=18).value = f"=IFERROR(AVERAGE(R{first_row_of_day_data}:R{last_row_of_day_data}),\"\")" # Q -> R
                ws.cell(row=row_num, column=20).value = f"=SUM(T{first_row_of_day_data}:T{last_row_of_day_data})" # S -> T
                ws.cell(row=row_num, column=21).value = f"=U{last_row_of_day_data}" # T -> U
                ws.cell(row=row_num, column=22).value = f"=V{last_row_of_day_data}" # U -> V
                ws.cell(row=row_num, column=23).value = f"=W{last_row_of_day_data}" # V -> W
                ws.cell(row=row_num, column=24).value = f"=X{last_row_of_day_data}" # W -> X
                ws.cell(row=row_num, column=25).value = f"=Y{last_row_of_day_data}" # X -> Y
                ws.cell(row=row_num, column=26).value = f"=Z{last_row_of_day_data}" # Y -> Z
                ws.cell(row=row_num, column=27).value = f"=AA{last_row_of_day_data}" # Z -> AA
            
            daily_summary_rows_info.append({
                "summary_row": row_num, "data_start_row": first_row_of_day_data,
                "data_end_row": last_row_of_day_data, "date_str": day_date_str
            })
            row_num += 1
            current_date += timedelta(days=1)

        if daily_summary_rows_info:
            weekly_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
            
            first_day_info = daily_summary_rows_info[0]
            date_cell_value_first_day = ws.cell(row=first_day_info["data_start_row"] - 1, column=1).value
            if date_cell_value_first_day and isinstance(date_cell_value_first_day, str) and " " in date_cell_value_first_day:
                first_day_date_str_for_range = date_cell_value_first_day.split(" ", 1)[1]
            else:
                first_day_date_str_for_range = first_day_info.get("date_str", "Unknown_Start")

            last_day_info = daily_summary_rows_info[-1]
            date_cell_value_last_day = ws.cell(row=last_day_info["summary_row"] - 1, column=1).value
            if date_cell_value_last_day and isinstance(date_cell_value_last_day, str) and " " in date_cell_value_last_day:
                last_day_date_str_for_range = date_cell_value_last_day.split(" ", 1)[1]
            else:
                last_day_date_str_for_range = last_day_info.get("date_str", "Unknown_End")
            
            week_range = f"{first_day_date_str_for_range} to {last_day_date_str_for_range}"

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            # Adjusted end_column for weekly summary merge (was 13, now 14)
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=14)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col); cell.fill = weekly_fill
                cell.font = Font(bold=True, color="FFFFFF", name="Dana"); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_num, column=1).value = f"Week {week + 1} ({week_range})"
            ws.row_dimensions[row_num].height = 30
            
            avg_sl_daily_cells = ",".join([f"O{info['summary_row']}" for info in daily_summary_rows_info if info['data_start_row'] <= info['data_end_row']]) # N -> O
            sum_risk_daily_cells = ",".join([f"P{info['summary_row']}" for info in daily_summary_rows_info if info['data_start_row'] <= info['data_end_row']]) # O -> P
            sum_target_r_daily_cells = ",".join([f"Q{info['summary_row']}" for info in daily_summary_rows_info if info['data_start_row'] <= info['data_end_row']]) # P -> Q
            avg_duration_daily_cells = ",".join([f"R{info['summary_row']}" for info in daily_summary_rows_info if info['data_start_row'] <= info['data_end_row']]) # Q -> R
            sum_result_daily_cells = ",".join([f"T{info['summary_row']}" for info in daily_summary_rows_info if info['data_start_row'] <= info['data_end_row']]) # S -> T

            if avg_sl_daily_cells: ws.cell(row=row_num, column=15).value = f"=IFERROR(AVERAGE({avg_sl_daily_cells}),\"\")" # 14 -> 15
            if sum_risk_daily_cells: ws.cell(row=row_num, column=16).value = f"=IFERROR(SUM({sum_risk_daily_cells}),\"\")" # 15 -> 16
            if sum_target_r_daily_cells: ws.cell(row=row_num, column=17).value = f"=IFERROR(SUM({sum_target_r_daily_cells}),\"\")" # 16 -> 17
            if avg_duration_daily_cells: ws.cell(row=row_num, column=18).value = f"=IFERROR(AVERAGE({avg_duration_daily_cells}),\"\")" # 17 -> 18
            if sum_result_daily_cells: ws.cell(row=row_num, column=20).value = f"=SUM({sum_result_daily_cells})" # 19 -> 20
            
            last_daily_summary_row = daily_summary_rows_info[-1]["summary_row"]
            ws.cell(row=row_num, column=21).value = f"=U{last_daily_summary_row}" # 20 -> 21
            ws.cell(row=row_num, column=22).value = f"=V{last_daily_summary_row}" # 21 -> 22
            ws.cell(row=row_num, column=23).value = f"=W{last_daily_summary_row}" # 22 -> 23
            ws.cell(row=row_num, column=24).value = f"=X{last_daily_summary_row}" # 23 -> 24
            ws.cell(row=row_num, column=25).value = f"=Y{last_daily_summary_row}" # 24 -> 25
            ws.cell(row=row_num, column=26).value = f"=Z{last_daily_summary_row}" # 25 -> 26
            ws.cell(row=row_num, column=27).value = f"=AA{last_daily_summary_row}" # 26 -> 27
            
            last_weekly_summary_row = row_num
            row_num += 1

    if first_data_row:
        monthly_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        # Adjusted end_column for monthly summary merge (was 13, now 14)
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=14)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col); cell.fill = monthly_fill
            cell.font = Font(bold=True, name="Dana"); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=row_num, column=1).value = "Monthly Summary"
        ws.row_dimensions[row_num].height = 30
        
        source_summary_row = last_weekly_summary_row if last_weekly_summary_row is not None else (daily_summary_rows_info[-1]["summary_row"] if daily_summary_rows_info else first_data_row -1)
        if source_summary_row < 2 : source_summary_row = row_num -1

        if source_summary_row >= first_data_row-1 :
            ws.cell(row=row_num, column=15).value = f"=O{source_summary_row}" # 14 -> 15
            ws.cell(row=row_num, column=16).value = f"=P{source_summary_row}" # 15 -> 16
            ws.cell(row=row_num, column=17).value = f"=Q{source_summary_row}" # 16 -> 17
            ws.cell(row=row_num, column=18).value = f"=R{source_summary_row}" # 17 -> 18
            ws.cell(row=row_num, column=20).value = f"=T{source_summary_row}" # 19 -> 20
            ws.cell(row=row_num, column=21).value = f"=U{source_summary_row}" # 20 -> 21
            ws.cell(row=row_num, column=22).value = f"=V{source_summary_row}" # 21 -> 22
            ws.cell(row=row_num, column=23).value = f"=W{source_summary_row}" # 22 -> 23
            ws.cell(row=row_num, column=24).value = f"=X{source_summary_row}" # 23 -> 24
            ws.cell(row=row_num, column=25).value = f"=Y{source_summary_row}" # 24 -> 25
            ws.cell(row=row_num, column=26).value = f"=Z{source_summary_row}" # 25 -> 26
            ws.cell(row=row_num, column=27).value = f"=AA{source_summary_row}" # 26 -> 27

    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    for r in all_data_rows:
        ws.conditional_formatting.add(f"P{r}", CellIsRule(operator='lessThan', formula=['1'], stopIfTrue=True, fill=green_fill)) # O -> P
        for col_letter in ['T', 'U', 'V']: # S, T, U -> T, U, V
            cell_ref = f"{col_letter}{r}"
            ws.conditional_formatting.add(cell_ref, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green_fill))
            ws.conditional_formatting.add(cell_ref, CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill))
        for col_letter in ['W', 'X', 'Y', 'Z']: # V, W, X, Y -> W, X, Y, Z
            cell_ref = f"{col_letter}{r}"
            ws.conditional_formatting.add(cell_ref, CellIsRule(operator='greaterThanOrEqual', formula=[str(initial_capital)], stopIfTrue=True, fill=green_fill))
            ws.conditional_formatting.add(cell_ref, CellIsRule(operator='lessThan', formula=[str(initial_capital)], stopIfTrue=True, fill=red_fill))
        cell_ref_aa = f"AA{r}" # Z -> AA
        ws.conditional_formatting.add(cell_ref_aa, CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add(cell_ref_aa, CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill))
        cell_ref_d = f"D{r}"
        ws.conditional_formatting.add(cell_ref_d, FormulaRule(formula=[f'OR(EXACT(D{r},"Strong"), EXACT(D{r},"Normal"), EXACT(D{r},"V-Strong"), EXACT(D{r},"V-Normal"))'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add(cell_ref_d, FormulaRule(formula=[f'OR(EXACT(D{r},"Weak"), EXACT(D{r},"V-Weak"))'], stopIfTrue=True, fill=red_fill))
        cell_ref_e = f"E{r}"
        ws.conditional_formatting.add(cell_ref_e, FormulaRule(formula=[f'AND(NOT(ISBLANK(E{r})),OR(EXACT(E{r},"Range"),EXACT(E{r},"V-Range")) )'], stopIfTrue=True, fill=red_fill))
        ws.conditional_formatting.add(cell_ref_e, FormulaRule(formula=[f'AND(NOT(ISBLANK(E{r})),NOT(OR(EXACT(E{r},"Range"),EXACT(E{r},"V-Range"))))'], stopIfTrue=True, fill=green_fill))
        cell_ref_f = f"F{r}"
        ws.conditional_formatting.add(cell_ref_f, FormulaRule(formula=[f'AND(NOT(ISBLANK(F{r})),OR(EXACT(F{r},"Increasing"),EXACT(F{r},"V-Increasing")) )'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add(cell_ref_f, FormulaRule(formula=[f'AND(NOT(ISBLANK(F{r})),NOT(OR(EXACT(F{r},"Increasing"),EXACT(F{r},"V-Increasing"))))'], stopIfTrue=True, fill=red_fill))

        # Conditional formatting for "Plan Adherence" (Column N, shifted from M)
        cell_ref_n = f"N{r}"
        ws.conditional_formatting.add(cell_ref_n, CellIsRule(operator='equal', formula=['"Full Adherence"'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add(cell_ref_n, FormulaRule(formula=[f'AND(N{r}<>"", N{r}<>"Full Adherence")'], stopIfTrue=True, fill=red_fill))
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in locals() or "__file__" in globals() else os.getcwd()
        output_dir = os.path.join(script_dir, 'output')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
        print(f"Successfully generated Excel file at:\n{os.path.abspath(output_path)}")
        open_file_os_agnostic(output_path) # Open the file
    except Exception as e:
        print(f"Error saving file: {e}")
        if "__file__" not in locals() and "__file__" not in globals():
            alt_output_path = os.path.join(os.getcwd(), filename)
            try:
                wb.save(alt_output_path)
                print(f"Successfully generated Excel file (fallback location):\n{os.path.abspath(alt_output_path)}")
                open_file_os_agnostic(alt_output_path) # Open the file from fallback location
            except Exception as e_alt:
                print(f"Error saving file (fallback attempt) or opening it: {e_alt}")

if __name__ == "__main__":
    try:
        date_input = input("Enter start date (YYYY-MM-DD) or press Enter for default (2025-04-28): ").strip()
        capital_input = input("Enter initial capital or press Enter for default (25000): ").strip()
        weeks_input = input("Enter number of weeks (1-4) or press Enter for default (4): ").strip()

        date_str = date_input if date_input else "2025-04-28"
        initial_capital = float(capital_input) if capital_input else 25000
        num_weeks_val = int(weeks_input) if weeks_input else 4

        if not 1 <= num_weeks_val <= 4:
            raise ValueError("Number of weeks must be between 1 and 4.")
        
        filename = f"trading_journal_{date_str}_to_{(datetime.strptime(date_str, '%Y-%m-%d').date() + timedelta(weeks=num_weeks_val)).strftime('%Y-%m-%d')}.xlsx"
        generate_trading_journal_excel(date_str, initial_capital, num_weeks_val, filename)

    except KeyboardInterrupt: print("\nOperation cancelled by user.")
    except ValueError as ve: print(f"Input Error: {ve}")
    except Exception as e: print(f"An unexpected error occurred: {e}")
    finally: sys.exit(0 if 'filename' in locals() else 1)
